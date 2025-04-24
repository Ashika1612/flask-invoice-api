import os
import time
import pandas as pd
from flask import Flask, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
import fitz  # PyMuPDF
import re
from openpyxl import load_workbook

app = Flask(__name__)

# Directories on the client's system
INPUT_FOLDER = '/path/to/input'  # Folder where clients upload Excel and PDF files
OUTPUT_FOLDER = '/path/to/output'  # Folder where processed CSV files will be saved
MASTER_FILE_PATH = '/path/to/master/MASTER_DATA_FILE.xlsx'  # Path to the master file
TEMPLATE_PATH = '/path/to/template/Template_File_Upload.xlsx'  # Path to the template file

# Allowed file extensions for upload (Excel and PDF)
ALLOWED_EXTENSIONS = {'xlsx', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_invoice_amount(pdf_path):
    """Extracts amount from 'Total Due USD' line in PDF"""
    doc = fitz.open(pdf_path)
    text = doc[0].get_text()
    doc.close()

    match = re.search(r"Total Due USD\s+([\d,]+\.\d{2})", text)
    if match:
        return float(match.group(1).replace(",", ""))
    raise ValueError(f"Amount not found in PDF. Text sample:\n{text[:500]}...")

def process_excel_file(input_file_path, template_path, output_folder, master_file_path):
    print(f"Started processing: {input_file_path}")

    # Try finding correct header row
    for skip in range(5):
        df = pd.read_excel(input_file_path, skiprows=skip)
        if 'L01 Material Price Group Key' in df.columns and 'Inv Net Amt' in df.columns and 'Material' in df.columns:
            break
    else:
        print("Required columns not found.")
        return

    df = df.dropna(subset=['L01 Material Price Group Key'])
    df.sort_values(by='L01 Material Price Group Key', inplace=True)

    # Extract total due value from corresponding PDF
    pdf_filename = os.path.splitext(os.path.basename(input_file_path))[0] + '.pdf'
    pdf_path = os.path.join(os.path.dirname(input_file_path), pdf_filename)

    if not os.path.exists(pdf_path):
        print(f"Corresponding PDF not found for: {input_file_path}")
        return

    try:
        total_due_value = extract_invoice_amount(pdf_path)
        print(f"Extracted total due from PDF: {total_due_value}")
    except Exception as e:
        print(f"Failed to extract invoice amount from PDF: {e}")
        return

    # Group and allocate values
    grouped = df.groupby('L01 Material Price Group Key').agg({'Inv Net Amt': 'sum'}).reset_index()
    grouped['Percentage'] = grouped['Inv Net Amt'] / grouped['Inv Net Amt'].sum()
    grouped['Allocated Amount'] = grouped['Percentage'] * total_due_value

    top_materials = df.sort_values(by='Inv Net Amt', ascending=False).groupby('L01 Material Price Group Key').first().reset_index()
    final_df = pd.merge(grouped, top_materials[['L01 Material Price Group Key', 'Material']], on='L01 Material Price Group Key')

    # Load template to get headers
    wb = load_workbook(template_path)
    ws = wb['Item Upload']
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    num_columns = len(headers)

    try:
        material_index = headers.index('Item_Number')
        amount_index = headers.index('Extended_Amount')
        upc_index = headers.index('UPC_Number')
        quantity_index = headers.index('Quantity')
    except ValueError:
        print("Required headers not found in template.")
        return

    # Prepare output data
    data_rows = []
    for _, r in final_df.iterrows():
        if pd.isna(r['Material']) or pd.isna(r['Allocated Amount']):
            continue
        row = [''] * num_columns
        row[material_index] = str(int(r['Material']))
        row[amount_index] = round(r['Allocated Amount'], 2)
        data_rows.append(row)

    # Read Master file and map UPC to Item Number
    try:
        master_df = pd.read_excel(master_file_path, usecols=[0, 1], names=['UPC', 'Item_Number'], header=0, dtype=str)
        master_df['Item_Number'] = master_df['Item_Number'].astype(str).str.strip()
        master_df['UPC'] = master_df['UPC'].astype(str).str.strip()
        master_df = master_df.drop_duplicates(subset='Item_Number', keep='first')
        upc_lookup = dict(zip(master_df['Item_Number'], master_df['UPC']))

        for row in data_rows:
            item = str(row[material_index]).strip()
            upc_value = upc_lookup.get(item, 'NA')

            if upc_value != 'NA':
                upc_value = upc_value.zfill(12)  # Ensure UPC is 12 digits
                row[upc_index] = f'="{upc_value}"'  # Force Excel to treat it as text
            else:
                row[upc_index] = 'NA'

        print("UPC mapping completed.")
    except Exception as e:
        print(f"Error reading master file: {e}")

    # Fill Quantity column with 1 only for actual data rows
    for row in data_rows:
        if row[material_index] not in [None, '', 'NA']:
            row[quantity_index] = 1
        else:
            row[quantity_index] = ''

    # Pad to 100 rows
    while len(data_rows) < 100:
        data_rows.append([''] * num_columns)

    # Create DataFrame and save CSV
    csv_df = pd.DataFrame(data_rows, columns=headers)
    output_filename = os.path.splitext(os.path.basename(input_file_path))[0] + '.csv'
    output_path = os.path.join(output_folder, output_filename)
    csv_df.to_csv(output_path, index=False)
    print(f"CSV Output saved: {output_path}")

@app.route('/process-invoice', methods=['POST'])
def process_invoice():
    if 'excel' not in request.files or 'pdf' not in request.files:
        return jsonify({"error": "Excel and PDF files are required!"}), 400

    excel_file = request.files['excel']
    pdf_file = request.files['pdf']

    # Secure the file names to prevent any security issues
    excel_filename = secure_filename(excel_file.filename)
    pdf_filename = secure_filename(pdf_file.filename)

    # Save the files temporarily to the input folder
    excel_path = os.path.join(INPUT_FOLDER, excel_filename)
    pdf_path = os.path.join(INPUT_FOLDER, pdf_filename)

    # Create the folders if they don't exist
    os.makedirs(INPUT_FOLDER, exist_ok=True)

    # Save the files to the server's input folder
    excel_file.save(excel_path)
    pdf_file.save(pdf_path)

    # Process the files
    try:
        process_excel_file(excel_path, TEMPLATE_PATH, OUTPUT_FOLDER, MASTER_FILE_PATH)
        return jsonify({"message": "Files processed successfully!"}), 200
    except Exception as e:
        return jsonify({"error": f"Error processing files: {str(e)}"}), 500

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(file_path):
        return send_from_directory(OUTPUT_FOLDER, filename, as_attachment=True)
    else:
        return jsonify({"error": "File not found!"}), 404
    
@app.route('/')
def home():
    return "✅ Flask app deployed successfully on Railway!"

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
